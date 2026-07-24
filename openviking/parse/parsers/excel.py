# Copyright (c) 2026 Beijing Volcano Engine Technology Co., Ltd.
# SPDX-License-Identifier: AGPL-3.0
"""
Excel (.xlsx/.xls/.xlsm) parser for OpenViking.

Converts Excel spreadsheets to Markdown then parses using MarkdownParser.
Inspired by microsoft/markitdown approach.
"""

import asyncio
import concurrent.futures
import time
from dataclasses import asdict, fields
from functools import partial
from multiprocessing import get_context
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openviking.parse.base import NodeType, ParseResult, ResourceNode, create_parse_result
from openviking.parse.parsers.base_parser import BaseParser
from openviking_cli.utils.config.parser_config import ExcelConfig, ParserConfig
from openviking_cli.utils.logger import get_logger

logger = get_logger(__name__)

_EXCEL_LAYOUT_EXECUTOR: Optional[concurrent.futures.ProcessPoolExecutor] = None
_EXCEL_LAYOUT_EXECUTOR_WORKERS: Optional[int] = None

# Non-configurable process-pool internals.
_EXCEL_PROCESS_POOL_START_METHOD = "spawn"
_EXCEL_PROCESS_POOL_MIN_BYTES = 200_000
_EXCEL_PROCESS_POOL_TIMEOUT_S = 120.0


def _get_excel_layout_executor(workers: int) -> concurrent.futures.ProcessPoolExecutor:
    global _EXCEL_LAYOUT_EXECUTOR, _EXCEL_LAYOUT_EXECUTOR_WORKERS
    workers = max(1, int(workers))
    if _EXCEL_LAYOUT_EXECUTOR is None:
        _EXCEL_LAYOUT_EXECUTOR = concurrent.futures.ProcessPoolExecutor(
            max_workers=workers,
            mp_context=get_context(_EXCEL_PROCESS_POOL_START_METHOD),
        )
        _EXCEL_LAYOUT_EXECUTOR_WORKERS = workers
        logger.info(
            f"[ExcelParserProcessPool] Started process pool "
            f"workers={workers} start_method={_EXCEL_PROCESS_POOL_START_METHOD}"
        )
    elif _EXCEL_LAYOUT_EXECUTOR_WORKERS != workers:
        logger.warning(
            "[ExcelParserProcessPool] Ignoring process_pool_workers=%s; "
            "pool already started with workers=%s",
            workers,
            _EXCEL_LAYOUT_EXECUTOR_WORKERS,
        )
    return _EXCEL_LAYOUT_EXECUTOR


def _build_excel_layout_in_process(
    *,
    path_str: str,
    temp_uri: str,
    instruction: str,
    layout_kwargs: Dict[str, Any],
    config_dict: Dict[str, Any],
    max_rows_per_sheet: int,
) -> Dict[str, Any]:
    """CPU-only child process worker.

    The worker must not touch VikingFS, DB, queues, or RequestContext. It only
    converts Excel to markdown and computes MarkdownParser layout ops.
    """
    import asyncio

    from openviking.parse.parsers.excel import ExcelParser
    from openviking_cli.utils.config.parser_config import ParserConfig

    started = time.perf_counter()
    path = Path(path_str)
    allowed_config_fields = {field.name for field in fields(ParserConfig)}
    filtered_config = {
        key: value for key, value in config_dict.items() if key in allowed_config_fields
    }
    parser = ExcelParser(
        config=ParserConfig.from_dict(filtered_config),
        max_rows_per_sheet=max_rows_per_sheet,
    )

    convert_started = time.perf_counter()
    if path.suffix.lower() == ".xls":
        markdown_content = parser._convert_xls_to_markdown(path)
    else:
        import openpyxl

        markdown_content = parser._convert_to_markdown(path, openpyxl)
    convert_s = time.perf_counter() - convert_started

    layout_started = time.perf_counter()
    layout = asyncio.run(
        parser._md_parser._compute_layout(
            markdown_content,
            temp_uri,
            source_path=str(path),
            instruction=instruction,
            **layout_kwargs,
        )
    )
    layout_s = time.perf_counter() - layout_started

    return {
        "layout": layout,
        "convert_s": convert_s,
        "layout_s": layout_s,
        "total_s": time.perf_counter() - started,
        "markdown_chars": len(markdown_content),
        "layout_ops": len(layout.ops),
        "layout_write_ops": sum(1 for op in layout.ops if op.kind == "write"),
        "layout_mkdir_ops": sum(1 for op in layout.ops if op.kind == "mkdir"),
    }


class ExcelParser(BaseParser):
    """
    Excel spreadsheet parser for OpenViking.

    Supports: .xlsx, .xls, .xlsm

    Converts Excel spreadsheets to Markdown using openpyxl,
    then delegates to MarkdownParser for tree structure creation.
    """

    def __init__(
        self, config: Optional[ParserConfig] = None, max_rows_per_sheet: int = 1000
    ):
        """
        Initialize Excel parser.

        Args:
            config: Parser configuration (prefer ExcelConfig for process-pool knobs)
            max_rows_per_sheet: Maximum rows to process per sheet (0 = unlimited)
        """
        from openviking.parse.parsers.markdown import MarkdownParser

        self._md_parser = MarkdownParser(config=config)
        self.config = config or ExcelConfig()
        self.max_rows_per_sheet = max_rows_per_sheet

    def _process_pool_enabled(self) -> bool:
        return bool(getattr(self.config, "enable_process_pool", False))

    def _process_pool_workers(self) -> int:
        return max(1, int(getattr(self.config, "process_pool_workers", 2) or 2))

    @property
    def supported_extensions(self) -> List[str]:
        return [".xlsx", ".xls", ".xlsm"]

    async def parse(self, source: Union[str, Path], instruction: str = "", **kwargs) -> ParseResult:
        """Parse Excel spreadsheet from file path."""
        path = Path(source)

        if path.exists():
            result = await self._parse_existing_path(path, instruction=instruction, **kwargs)
        else:
            result = await self._md_parser.parse_content(
                str(source), instruction=instruction, **kwargs
            )
        result.source_format = path.suffix.lstrip(".") if path.exists() else "xlsx"
        result.parser_name = "ExcelParser"
        return result

    async def _parse_existing_path(
        self, path: Path, instruction: str = "", **kwargs
    ) -> ParseResult:
        if self._should_use_process_pool(path, kwargs):
            try:
                return await self._parse_existing_path_process_pool(
                    path, instruction=instruction, **kwargs
                )
            except Exception as exc:
                logger.warning(
                    f"[ExcelParserProcessPool] Falling back to in-process parse: {exc}",
                    exc_info=True,
                )

        # Use xlrd for legacy .xls, openpyxl for .xlsx/.xlsm
        if path.suffix.lower() == ".xls":
            markdown_content = await asyncio.to_thread(self._convert_xls_to_markdown, path)
        else:
            import openpyxl

            markdown_content = await asyncio.to_thread(self._convert_to_markdown, path, openpyxl)
        return await self._md_parser.parse_content(
            markdown_content, source_path=str(path), instruction=instruction, **kwargs
        )

    def _should_use_process_pool(self, path: Path, kwargs: Dict[str, Any]) -> bool:
        if not self._process_pool_enabled():
            return False
        if path.suffix.lower() == ".xls":
            return False
        if kwargs.get("enable_link_rewrite") or kwargs.get("base_dir") or kwargs.get("allowed_media_dirs"):
            logger.debug("[ExcelParserProcessPool] Disabled for link/media rewrite parse")
            return False
        try:
            return path.stat().st_size >= _EXCEL_PROCESS_POOL_MIN_BYTES
        except OSError:
            return False

    async def _parse_existing_path_process_pool(
        self, path: Path, instruction: str = "", **kwargs
    ) -> ParseResult:
        parse_started = time.perf_counter()
        loop = asyncio.get_running_loop()
        temp_uri = self._md_parser._create_temp_uri()
        layout_kwargs = {
            key: value
            for key, value in kwargs.items()
            if key in {"resource_name", "source_name"} and isinstance(value, str)
        }
        future = loop.run_in_executor(
            _get_excel_layout_executor(self._process_pool_workers()),
            partial(
                _build_excel_layout_in_process,
                path_str=str(path),
                temp_uri=temp_uri,
                instruction=instruction,
                layout_kwargs=layout_kwargs,
                config_dict=asdict(self.config),
                max_rows_per_sheet=self.max_rows_per_sheet,
            ),
        )

        worker_started = time.perf_counter()
        worker_result = await asyncio.wait_for(
            future, timeout=_EXCEL_PROCESS_POOL_TIMEOUT_S
        )
        worker_s = time.perf_counter() - worker_started
        layout = worker_result["layout"]

        self._md_parser._rewrite_ctx = {
            "enabled": False,
            "source_path": str(path),
            "doc_name": layout.doc_name,
            "root_dir": layout.root_dir,
            "import_root": None,
            "base_dir": None,
            "allowed_media_dirs": None,
        }
        try:
            apply_started = time.perf_counter()
            await self._md_parser._apply_layout(layout)
            apply_s = time.perf_counter() - apply_started
        finally:
            self._md_parser._rewrite_ctx = None

        parse_time = time.perf_counter() - parse_started
        logger.info(
            f"[ExcelParserProcessPool] path={path} total={parse_time:.3f}s "
            f"worker_wall={worker_s:.3f}s convert={worker_result.get('convert_s', -1.0):.3f}s "
            f"layout={worker_result.get('layout_s', -1.0):.3f}s apply={apply_s:.3f}s "
            f"chars={worker_result.get('markdown_chars')} ops={worker_result.get('layout_ops')} "
            f"writes={worker_result.get('layout_write_ops')} mkdirs={worker_result.get('layout_mkdir_ops')}"
        )

        root = ResourceNode(
            type=NodeType.ROOT,
            title=layout.doc_title,
            level=0,
            meta=layout.meta.get("frontmatter", {}),
        )
        result = create_parse_result(
            root=root,
            source_path=str(path),
            source_format="markdown",
            parser_name="MarkdownParser",
            parse_time=parse_time,
            meta=layout.meta,
            warnings=layout.warnings,
        )
        result.temp_dir_path = layout.temp_uri
        return result

    async def parse_content(
        self, content: str, source_path: Optional[str] = None, instruction: str = "", **kwargs
    ) -> ParseResult:
        """Parse content - delegates to MarkdownParser."""
        result = await self._md_parser.parse_content(content, source_path, **kwargs)
        result.source_format = "xlsx"
        result.parser_name = "ExcelParser"
        return result

    def _convert_xls_to_markdown(self, path: Path) -> str:
        """Convert legacy .xls spreadsheet to Markdown using xlrd."""
        import xlrd

        # formatting_info=True enables xlrd to detect date cells via XL_CELL_DATE
        # instead of reporting them as XL_CELL_NUMBER with raw float serials
        wb = xlrd.open_workbook(str(path), formatting_info=True, on_demand=True)
        try:
            return self._build_xls_markdown(wb, path, xlrd)
        finally:
            wb.release_resources()

    def _build_xls_markdown(self, wb, path: Path, xlrd) -> str:
        """Build markdown from xlrd workbook."""
        markdown_parts = []
        markdown_parts.append(f"# {path.stem}")
        markdown_parts.append(f"**Sheets:** {wb.nsheets}")

        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            parts = [f"## Sheet: {sheet.name}"]

            if sheet.nrows == 0 or sheet.ncols == 0:
                parts.append("*Empty sheet*")
                markdown_parts.append("\n\n".join(parts))
                continue

            parts.append(f"**Dimensions:** {sheet.nrows} rows × {sheet.ncols} columns")

            rows_to_process = sheet.nrows
            if self.max_rows_per_sheet > 0:
                rows_to_process = min(sheet.nrows, self.max_rows_per_sheet)

            rows = []
            for row_idx in range(rows_to_process):
                row_data = []
                for col_idx in range(sheet.ncols):
                    row_data.append(self._format_xls_cell(sheet.cell(row_idx, col_idx), wb, xlrd))
                rows.append(row_data)

            if rows:
                from openviking.parse.base import format_table_to_markdown

                parts.append(format_table_to_markdown(rows, has_header=True))

            if self.max_rows_per_sheet > 0 and sheet.nrows > self.max_rows_per_sheet:
                parts.append(
                    f"\n*... {sheet.nrows - self.max_rows_per_sheet} more rows truncated ...*"
                )

            markdown_parts.append("\n\n".join(parts))

        return "\n\n".join(markdown_parts)

    @staticmethod
    def _format_xls_cell(cell, wb, xlrd) -> str:
        """Format a single xlrd cell value with proper type handling."""
        if cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
            return ""
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                dt = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                # Include time component if non-zero
                if dt[3] or dt[4] or dt[5]:
                    return (
                        f"{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d} {dt[3]:02d}:{dt[4]:02d}:{dt[5]:02d}"
                    )
                return f"{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d}"
            except Exception:
                return str(cell.value)
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return "TRUE" if cell.value else "FALSE"
        if cell.ctype == xlrd.XL_CELL_ERROR:
            # xlrd error code map
            error_map = {
                0x00: "#NULL!",
                0x07: "#DIV/0!",
                0x0F: "#VALUE!",
                0x17: "#REF!",
                0x1D: "#NAME?",
                0x24: "#NUM!",
                0x2A: "#N/A",
            }
            return error_map.get(cell.value, f"#ERR({cell.value})")
        if cell.ctype == xlrd.XL_CELL_NUMBER:
            # Display integers without trailing .0
            if cell.value == int(cell.value):
                return str(int(cell.value))
            return str(cell.value)
        # XL_CELL_TEXT or fallback
        return str(cell.value) if cell.value is not None else ""

    def _convert_to_markdown(self, path: Path, openpyxl) -> str:
        """Convert Excel spreadsheet to Markdown string."""
        wb = openpyxl.load_workbook(path, data_only=True)

        markdown_parts = []
        markdown_parts.append(f"# {path.stem}")
        markdown_parts.append(f"**Sheets:** {len(wb.sheetnames)}")

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_content = self._convert_sheet(sheet, sheet_name)
            markdown_parts.append(sheet_content)

        return "\n\n".join(markdown_parts)

    def _convert_sheet(self, sheet, sheet_name: str) -> str:
        """Convert a single sheet to markdown."""
        parts = []
        parts.append(f"## Sheet: {sheet_name}")

        max_row = sheet.max_row
        max_col = sheet.max_column

        if max_row == 0 or max_col == 0:
            parts.append("*Empty sheet*")
            return "\n\n".join(parts)

        parts.append(f"**Dimensions:** {max_row} rows × {max_col} columns")

        rows_to_process = max_row
        if self.max_rows_per_sheet > 0:
            rows_to_process = min(max_row, self.max_rows_per_sheet)

        rows = []
        for _row_idx, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=rows_to_process, values_only=True), 1
        ):
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                else:
                    row_data.append(str(cell))
            rows.append(row_data)

        if rows:
            from openviking.parse.base import format_table_to_markdown

            table_md = format_table_to_markdown(rows, has_header=True)
            parts.append(table_md)

        if self.max_rows_per_sheet > 0 and max_row > self.max_rows_per_sheet:
            parts.append(f"\n*... {max_row - self.max_rows_per_sheet} more rows truncated ...*")

        return "\n\n".join(parts)
